import hashlib
import os
import shutil
import subprocess
import tempfile
import threading
import time
from typing import List, Optional

import keyboard
import pyautogui
import pyperclip

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

_image_cache = {}


def _click_image_if_visible(image_path: str, click_times: int = 1, stop_event: Optional[threading.Event] = None) -> bool:
    """单次检测屏幕上的图片，如果找到则点击并返回 True。"""
    if not image_path:
        return False
    if stop_event and stop_event.is_set():
        return False
    safe_path = _get_ascii_safe_path(image_path)
    try:
        location = pyautogui.locateOnScreen(safe_path, confidence=0.9)
    except Exception as exc:
        print(f"图像检测异常 {safe_path}: {exc}")
        return False
    if not location:
        return False
    center = pyautogui.center(location)
    actual_clicks = max(int(click_times) if click_times else 1, 1)
    for _ in range(actual_clicks):
        if stop_event and stop_event.is_set():
            return False
        pyautogui.click(center)
        time.sleep(0.1)
    print(f"检测到图片 {safe_path}，已点击 {actual_clicks} 次")
    return True


def _get_ascii_safe_path(image_path: str) -> str:
    """
    OpenCV 在 Windows 下无法读取包含非 ASCII 字符的路径。
    当检测到路径包含中文等字符时，将图片复制到临时目录的 ASCII 路径后返回。
    """
    if not image_path:
        return image_path

    abs_path = os.path.abspath(image_path)

    if abs_path.isascii():
        return abs_path

    global _image_cache
    try:
        src_mtime = os.path.getmtime(abs_path)
    except OSError:
        return abs_path
    cached = _image_cache.get(abs_path)
    if cached:
        cached_path, cached_mtime = cached
        if os.path.exists(cached_path) and cached_mtime == src_mtime:
            return cached_path

    temp_dir = os.path.join(tempfile.gettempdir(), "auto_input_images")
    os.makedirs(temp_dir, exist_ok=True)
    suffix = os.path.splitext(abs_path)[1]
    hashed_name = hashlib.md5(abs_path.encode("utf-8")).hexdigest()
    temp_path = os.path.join(temp_dir, f"{hashed_name}{suffix}")

    shutil.copy2(abs_path, temp_path)
    _image_cache[abs_path] = (temp_path, src_mtime)
    return temp_path

# 操作函数
def input_text(text, max_retry: int = 3):
    """将文本写入剪贴板后粘贴，失败时自动重试并降级为逐字输入。"""
    normalized = text if isinstance(text, str) else str(text)

    for attempt in range(1, max_retry + 1):
        try:
            pyperclip.copy(normalized)
            time.sleep(0.05)  # 等待剪贴板可用
            if pyperclip.paste() != normalized:
                raise ValueError("剪贴板内容与目标不一致")

            keyboard.press_and_release('ctrl+v')
            time.sleep(0.2)
            print(f"粘贴成功（第 {attempt} 次尝试）: {normalized}")
            return True
        except Exception as exc:
            print(f"粘贴失败（第 {attempt} 次尝试）: {exc}")
            time.sleep(0.3)

    # 降级方案：逐字符输入，避免流程完全中断
    print("多次粘贴失败，切换为逐字符输入模式")
    pyautogui.typewrite(normalized, interval=0.02)
    time.sleep(0.2)
    return True

def find_and_click(
    image_path,
    d_x=0,
    d_y=0,
    click_times=0,
    stop_event: Optional[threading.Event] = None,
    retry_interval: float = 1.0,
    timeout: Optional[float] = None
):
    """
    在屏幕上查找图片并点击。默认行为为无限重试，除非提供 `timeout`（秒）或 `stop_event` 被设置。
    Args:
        timeout: 如果为数字，表示在该秒数后放弃并返回 False；如果为 None，则保持原先的无限重试行为。
    """
    start_time = time.time()
    while True:
        # 检查外部停止信号
        if stop_event and stop_event.is_set():
            print("find_and_click: 收到停止信号，中止查找。")
            return False

        # 检查超时条件（如果启用）
        if timeout is not None and timeout >= 0:
            elapsed = time.time() - start_time
            if elapsed >= float(timeout):
                print(f"find_and_click: 等待图片超时 {timeout} 秒，跳过 -> {image_path}")
                return False

        safe_path = image_path
        try:
            safe_path = _get_ascii_safe_path(image_path)
            location = pyautogui.locateOnScreen(safe_path, confidence=0.9)
            if location:
                center = pyautogui.center(location)
                click_position = (center.x + d_x, center.y + d_y)
                for _ in range(click_times):
                    pyautogui.click(click_position)
                time.sleep(1)
                print(f"图像识别: 横轴平移{d_x}, 纵轴平移{d_y}, 成功点击 {safe_path} ,点击{click_times}次")
                return True
            else:
                print(f"未找到目标 {safe_path}，重试中...")
        except Exception as e:
            print(f"未找到目标 {safe_path}，重试中...")

        time.sleep(max(retry_interval, 0.1))

def move_and_click(
    X,
    Y,
    scroll_times,
    scroll_distance=0,
    click_times=0,
    image_path: Optional[str] = None,
    stop_event: Optional[threading.Event] = None
):
    """移动/滚动并尝试在过程中识别图片。"""
    try:
        pos = pyautogui.position()
        target_x, target_y = pos.x + X, pos.y + Y
        pyautogui.moveTo(target_x, target_y, duration=0.2)
        print(f"鼠标移动: 当前({pos.x}, {pos.y}) -> 目标({target_x}, {target_y})")

        if image_path and _click_image_if_visible(image_path, click_times, stop_event):
            return True

        for _ in range(scroll_times):
            if stop_event and stop_event.is_set():
                print("move_and_click: 收到停止信号，结束滚动。")
                return False
            pyautogui.scroll(-scroll_distance)
            print(f"滚动: 次数+1, 距离 {scroll_distance}")
            time.sleep(0.3)
            if image_path and _click_image_if_visible(image_path, click_times, stop_event):
                return True

        for _ in range(click_times):
            if stop_event and stop_event.is_set():
                print("move_and_click: 收到停止信号，终止点击。")
                return False
            pyautogui.click()
            time.sleep(0.1)

        if image_path:
            print(f"未在滚动过程中检测到图片 {image_path}，已完成默认点击。")
        else:
            print(f"鼠标移动: 横轴平移{X}, 纵轴平移{Y}, 滚动次数{scroll_times}, 滚动距离{scroll_distance},点击{click_times}次")
        return True
    except Exception as e:
        print(f"执行过程中发生错误: {e}")
        return False
# move_and_click(X=-150, Y= 0,click_time=1)


def wait(seconds: float = 1.0):
    """简单的等待函数，支持浮点秒数。"""
    try:
        duration = max(float(seconds), 0.0)
    except (TypeError, ValueError):
        duration = 0.0

    print(f"等待 {duration} 秒...")
    time.sleep(duration)

def simulate_key(keys: str):
    """
    模拟按键操作，支持单个按键和组合键。
    
    Args:
        keys: 按键字符串，支持以下格式：
            - 单个按键: "A", "enter", "space", "tab" 等
            - 组合键: "ctrl+A", "ctrl+shift+A", "alt+tab" 等
            - 多个按键（空格分隔）: "ctrl+A space enter" 会依次执行三个操作
    
    示例:
        simulate_key("A")                    # 按 A 键
        simulate_key("ctrl+A")               # 按 Ctrl+A
        simulate_key("ctrl+shift+A")         # 按 Ctrl+Shift+A
        simulate_key("ctrl+A space enter")   # 依次执行 Ctrl+A, Space, Enter
    """
    if not keys or not keys.strip():
        print("按键字符串为空，跳过执行")
        return
    
    # 按键名称映射（将常见名称转换为 keyboard 库识别的格式）
    key_mapping = {
        'ctrl': 'ctrl',
        'control': 'ctrl',
        'alt': 'alt',
        'shift': 'shift',
        'win': 'win',
        'windows': 'win',
        'cmd': 'cmd',
        'command': 'cmd',
        'space': 'space',
        'enter': 'enter',
        'return': 'enter',
        'tab': 'tab',
        'esc': 'esc',
        'escape': 'esc',
        'backspace': 'backspace',
        'delete': 'delete',
        'del': 'delete',
        'up': 'up',
        'down': 'down',
        'left': 'left',
        'right': 'right',
        'home': 'home',
        'end': 'end',
        'pageup': 'page up',
        'pagedown': 'page down',
        'f1': 'f1', 'f2': 'f2', 'f3': 'f3', 'f4': 'f4',
        'f5': 'f5', 'f6': 'f6', 'f7': 'f7', 'f8': 'f8',
        'f9': 'f9', 'f10': 'f10', 'f11': 'f11', 'f12': 'f12',
    }
    
    def normalize_key(key: str) -> str:
        """标准化按键名称"""
        key_lower = key.lower().strip()
        return key_mapping.get(key_lower, key_lower)
    
    def parse_and_press(key_sequence: str):
        """解析并按下单个按键或组合键"""
        key_sequence = key_sequence.strip()
        if not key_sequence:
            return
        
        # 检查是否是组合键（包含 +）
        if '+' in key_sequence:
            # 解析组合键
            parts = [normalize_key(part) for part in key_sequence.split('+')]
            modifiers = []
            main_key = None
            for part in parts:
                if part in ['ctrl', 'alt', 'shift', 'win', 'cmd']:
                    modifiers.append(part)
                else:
                    main_key = part
            
            if not main_key:
                print(f"组合键解析失败: {key_sequence}，缺少主键")
                return
            
            # 按下修饰键
            for mod in modifiers:
                keyboard.press(mod)
            time.sleep(0.05)  # 短暂延迟确保修饰键已按下
            
            # 按下主键
            keyboard.press_and_release(main_key)
            time.sleep(0.05)
            
            # 释放修饰键
            for mod in reversed(modifiers):
                keyboard.release(mod)
            
            print(f"模拟按键: {key_sequence}")
        else:
            # 单个按键
            normalized = normalize_key(key_sequence)
            try:
                keyboard.press_and_release(normalized)
                print(f"模拟按键: {key_sequence}")
                time.sleep(0.05)
            except Exception as e:
                print(f"模拟按键失败 '{key_sequence}': {e}")
    
    # 按空格分割多个按键操作
    key_sequences = keys.strip().split()
    
    for key_seq in key_sequences:
        try:
            parse_and_press(key_seq)
            time.sleep(0.1)  # 按键之间的间隔
        except Exception as e:
            print(f"模拟按键失败 '{key_seq}': {e}")
            continue

def read_excel_first_column(file_path: str):
    """
    读取 Excel 文件第一列的非空内容，保持原有顺序。
    """
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl 依赖，请先安装 openpyxl")

    if not file_path:
        raise ValueError("Excel 文件路径为空")

    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"Excel 文件不存在: {abs_path}")

    workbook = load_workbook(abs_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = []
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            cell_value = row[0]
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            if text:
                values.append(text)
    finally:
        workbook.close()

    if not values:
        raise ValueError(f"Excel 文件第一列没有可用数据: {abs_path}")
    return values

def run_python(environment_path: str, program_path: str, args: Optional[List[str]] = None):
    """
    使用指定 Python 解释器运行目标 .py 脚本。

    Args:
        environment_path: Python 解释器路径（一般为 venv/Scripts/python.exe）。
        program_path:     需要执行的 Python 脚本路径。
        args:             传递给脚本的额外命令行参数列表。
    """
    interpreter_path = (environment_path or "").strip()
    script_path = (program_path or "").strip()

    if not interpreter_path:
        raise ValueError("Python 解释器路径不能为空")
    if not script_path:
        raise ValueError("脚本路径不能为空")

    interpreter_abs = os.path.normpath(os.path.abspath(interpreter_path))
    script_abs = os.path.normpath(os.path.abspath(script_path))

    if not os.path.isfile(interpreter_abs):
        raise FileNotFoundError(f"Python 解释器不存在: {interpreter_abs}")
    if not os.path.exists(script_abs):
        raise FileNotFoundError(f"脚本文件不存在: {script_abs}")

    working_dir = os.path.dirname(script_abs) or None

    try:
        extra_args = [str(arg) for arg in args] if args else []
        process = subprocess.Popen(
            [interpreter_abs, script_abs, *extra_args],
            cwd=working_dir
        )
        print(
            f"run_python: 使用 {interpreter_abs} 运行 {script_abs} "
            f"参数 {extra_args or '[]'}，工作目录 {working_dir or os.getcwd()}"
        )
        return process
    except Exception as exc:
        print(f"run_python: 启动失败 -> {exc}")
        raise

def if_image_condition_check(image_path_1, image_path_2):
    """
    检查屏幕上是否存在指定的两个图片之一，并返回相应的分支。

    Args:
        image_path_1: 第一个要检查的图片路径。
        image_path_2: 第二个要检查的图片路径。

    Returns:
        str:  如果识别到 image_path_1，返回 "image1_output"。
              如果识别到 image_path_2，返回 "image2_output"。
              如果两个图片都未找到，返回 "not_found_output"。
    """
    safe_image_1 = _get_ascii_safe_path(image_path_1)
    safe_image_2 = _get_ascii_safe_path(image_path_2)

    print(f"IF 条件检查函数执行: 检查 图片1: {safe_image_1}, 图片2: {safe_image_2}") # 调试信息

    location_1 = pyautogui.locateOnScreen(safe_image_1, confidence=0.9) # 识别 图片1
    if location_1:
        print(f"IF 条件检查函数: 识别到 图片1: {safe_image_1}") # 调试信息
        return "image1_output" # 返回 "image1_output" 分支

    location_2 = pyautogui.locateOnScreen(safe_image_2, confidence=0.9) # 识别 图片2
    if location_2:
        print(f"IF 条件检查函数: 识别到 图片2: {safe_image_2}") # 调试信息
        return "image2_output" # 返回 "image2_output" 分支

    print(f"IF 条件检查函数: 未找到 图片1: {safe_image_1} 和 图片2: {safe_image_2}") # 调试信息
    return "not_found_output" # 如果两个图片都未找到，返回 "not_found_output" 分支
